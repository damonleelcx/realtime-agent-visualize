"""Comparison renderers: self-contained HTML + valid, openable Office files."""

from __future__ import annotations

import io

from agent.comparison import run_comparison
from agent.skills.compare_office import to_docx, to_pptx, to_xlsx
from agent.skills.compare_viz import render_comparison_html
from tests.comparison_fixtures import FROZEN, market_data_stub


def _result(title="Gold vs Bitcoin"):
    return run_comparison(
        ["GC=F", "BTC-USD"], "2023-01-02", "2023-06-01", [],  # no artifacts written
        title=title, market_data_fn=market_data_stub(), clock=lambda: FROZEN,
    ).result


def test_html_is_self_contained_and_interactive() -> None:
    html = render_comparison_html(_result())
    assert "echarts.init" in html                          # chart wired
    assert "window.__CMP__" in html                        # data embedded
    # no external resources: no CDN <script src> / <link href> to the network
    assert "src=\"http" not in html and "href=\"http" not in html.split("<ul")[0]
    assert "cdn" not in html.lower()


def test_html_escapes_injection_in_title() -> None:
    html = render_comparison_html(_result(title="<script>alert(1)</script>PWN"))
    assert "<script>alert(1)</script>PWN" not in html      # raw injection neutralized
    assert "alert(1)" in html                              # ...but present, escaped
    assert "&lt;script&gt;alert(1)" in html


def test_xlsx_has_backtest_workbook_tabs() -> None:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(to_xlsx(_result())))
    assert wb.sheetnames == ["Overview", "Prices", "Returns", "Correlation",
                             "Backtests", "Sources"]
    # Sources tab carries a clickable provenance URL for each asset
    urls = [row[2] for row in wb["Sources"].iter_rows(min_row=2, values_only=True)]
    assert all(str(u).startswith("http") for u in urls)


def test_pptx_and_docx_open() -> None:
    from docx import Document
    from pptx import Presentation

    prs = Presentation(io.BytesIO(to_pptx(_result())))
    assert sum(1 for _ in prs.slides) >= 5
    doc = Document(io.BytesIO(to_docx(_result())))
    text = "\n".join(p.text for p in doc.paragraphs)
    assert "Methodology" in text and "252-day" in text
